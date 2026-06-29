#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_dashboard.py
-------------------
讀取瓦斯用量 Excel (.xlsm) 的 Daily_Summary sheet，更新 index.html 的每日用量資料；
互動式詢問是否要新增一筆從 gas.e-letter.com.tw 抄錄的帳單（度數/單價/合計）；
然後 git commit + push 到 GitHub。

雙擊執行：更新瓦斯Dashboard.bat

使用方式：
  python update_dashboard.py
  python update_dashboard.py --dry-run
  python update_dashboard.py --no-bill-prompt   (略過詢問新帳單，純同步用量)
"""

import argparse, json, re, subprocess, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，請執行：pip install openpyxl")
    sys.exit(1)

DEFAULT_EXCEL = r"D:\AI application code\gas-bill-analysis\Gas usage.xlsm"
HTML_FILE     = "index.html"
SHEET         = "Daily_Summary"

# 實際帳單記錄 — 從 gas.e-letter.com.tw 歷史帳單手動抄錄，新帳單請手動新增
KNOWN_BILLS = [
    {"deg": 1, "rate": 13.14, "base": 300, "total": 313},
    {"deg": 1, "rate": 13.14, "base": 300, "total": 313},
    {"deg": 2, "rate": 13.28, "base": 300, "total": 327},
    {"deg": 1, "rate": 13.41, "base": 300, "total": 313},
]


def read_daily(path: Path) -> list:
    print(f"讀取 {path.name} ...")
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"找不到工作表「{SHEET}」"); sys.exit(1)
    ws = wb[SHEET]
    rows = []
    for date, year, quarter, month, day, usage in ws.iter_rows(min_row=2, values_only=True):
        if date is None:
            continue
        rows.append([date.strftime("%Y-%m-%d"), round(float(usage), 3) if usage is not None else 0])
    wb.close()
    print(f"讀取完成，共 {len(rows)} 天記錄")
    return rows


def update_html(html_path: Path, daily: list, bills: list) -> bool:
    content = html_path.read_text(encoding="utf-8")
    daily_json = json.dumps(daily, ensure_ascii=False, separators=(",", ":"))
    bills_json = json.dumps(bills, ensure_ascii=False)
    latest_rate = bills[-1]["rate"] if bills else 13.0

    pattern = r"const DAILY = .*?;\n// 實際帳單記錄.*?\nconst KNOWN_BILLS = .*?;"
    replacement = (
        f"const DAILY = {daily_json};\n"
        f"// 實際帳單記錄（手動從 gas.e-letter.com.tw 抄錄，無法自動同步）\n"
        f"const KNOWN_BILLS = {bills_json};"
    )
    new_content, count = re.subn(pattern, replacement, content, flags=re.DOTALL)
    if count == 0:
        print("找不到 DATA BLOCK 標記"); return False

    new_content = new_content.replace("{{LATEST_RATE}}", str(latest_rate))

    html_path.write_text(new_content, encoding="utf-8")
    print(f"index.html 已更新（{len(daily)} 天 / {len(bills)} 筆帳單）")
    return True


def prompt_new_bill():
    print()
    ans = input("是否要新增一筆 gas.e-letter.com.tw 的帳單記錄？(y/n)：").strip().lower()
    if ans != "y":
        return None
    try:
        deg = float(input("  計費氣量（度數）：").strip())
        rate = float(input("  單價：").strip())
        base = float((input("  基本費（直接按 Enter 預設 300）：").strip() or "300"))
        total = round(deg * rate + base)
        confirm_total = input(f"  合計金額（直接按 Enter 採用試算值 ${total}）：").strip()
        if confirm_total:
            total = float(confirm_total)
        return {"deg": deg, "rate": rate, "base": base, "total": total}
    except ValueError:
        print("輸入格式錯誤，已取消新增帳單。")
        return None


def save_known_bills(py_path: Path, bills: list):
    content = py_path.read_text(encoding="utf-8")
    bills_src = "[\n" + "\n".join(
        f'    {{"deg": {b["deg"]}, "rate": {b["rate"]}, "base": {b["base"]}, "total": {b["total"]}}},'
        for b in bills
    ) + "\n]"
    pattern = r"KNOWN_BILLS = \[.*?\]"
    new_content, count = re.subn(pattern, f"KNOWN_BILLS = {bills_src}", content, count=1, flags=re.DOTALL)
    if count == 0:
        print("找不到 KNOWN_BILLS 區塊，無法寫回腳本"); return False
    py_path.write_text(new_content, encoding="utf-8")
    return True


def git_push(message: str, files: list):
    for f in files:
        subprocess.run(["git", "add", f], check=True)
    result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
    if "nothing to commit" in result.stdout + result.stderr:
        print("資料未變動，不需要 push"); return
    if result.returncode != 0:
        print(f"commit 失敗：{result.stderr}"); sys.exit(1)
    subprocess.run(["git", "push"], check=True)
    print("已成功 push 到 GitHub！")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--message", default="")
    parser.add_argument("--no-bill-prompt", action="store_true")
    args = parser.parse_args()

    repo_root = Path(__file__).parent
    excel_path = Path(args.excel) if Path(args.excel).is_absolute() else repo_root / args.excel
    html_path = repo_root / HTML_FILE
    py_path = Path(__file__)

    if not excel_path.exists():
        print(f"找不到 Excel：{excel_path}"); sys.exit(1)

    daily = read_daily(excel_path)
    if not daily:
        print("沒有讀到資料"); sys.exit(1)

    bills = KNOWN_BILLS
    if not args.no_bill_prompt:
        new_bill = prompt_new_bill()
        if new_bill:
            bills = KNOWN_BILLS + [new_bill]
            if save_known_bills(py_path, bills):
                print(f"  已新增帳單並寫回 {py_path.name}：{new_bill}")
            files_changed = [HTML_FILE, py_path.name]
        else:
            files_changed = [HTML_FILE]
    else:
        files_changed = [HTML_FILE]

    if not update_html(html_path, daily, bills):
        sys.exit(1)

    if args.dry_run:
        print("dry-run，跳過 push"); return

    latest = daily[-1][0]
    msg = args.message or f"update dashboard: 最新資料至 {latest}"
    git_push(msg, files_changed)
    print(f"完成！Dashboard 已更新至 {latest}")


if __name__ == "__main__":
    main()
